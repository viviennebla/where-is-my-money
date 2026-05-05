import csv
import io
from openpyxl import load_workbook

HEADER_KEYWORDS = [
    "交易时间", "交易日期", "时间", "日期",
    "金额", "收入", "支出", "收支", "收/支",
    "商户", "交易对方", "对方", "商户名称",
    "商品", "描述", "备注", "商品描述",
    "订单号", "交易单号", "商户单号",
    "支付方式", "付款方式",
    "币种", "货币",
    "当前状态", "交易类型",
]


def find_header_row(rows: list[list[str]]) -> int:
    """Find the real header row by scoring each row against known keywords."""
    best_score = 0
    best_idx = 0
    for i, row in enumerate(rows):
        if not row:
            continue
        score = 0
        for cell in row:
            cell_str = str(cell).strip()
            if not cell_str:
                continue
            for kw in HEADER_KEYWORDS:
                if kw in cell_str:
                    score += 1
                    break
        # Bonus for rows with more non-empty cells
        non_empty = sum(1 for c in row if str(c).strip())
        score += min(non_empty, 5)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def parse_file_data(
    filename: str,
    content: bytes,
    header_row_index: int = 0,
    row_limit: int = 0,
) -> tuple[list[str], list[list[str]]]:
    """Parse file starting from header_row_index. Returns (headers, data_rows)."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return _parse_excel_data(content, header_row_index, row_limit)
    else:
        return _parse_csv_data(content, header_row_index, row_limit)


def parse_file_preview(
    filename: str,
    content: bytes,
    header_row_index: int | None = None,
    preview_rows: int = 20,
) -> dict:
    """
    Parse file for preview. If header_row_index is None, auto-detect.
    Returns {"headers": [...], "sample_rows": [[...], ...], "header_row_index": int, "total_rows": int}.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        all_rows = _read_excel_rows(content)
    else:
        all_rows = _read_csv_rows(content)

    total_rows = len(all_rows)

    if header_row_index is None:
        header_row_index = find_header_row(all_rows)

    headers = [str(c).strip() for c in all_rows[header_row_index]] if header_row_index < len(all_rows) else []

    sample_start = header_row_index + 1
    sample_end = min(sample_start + preview_rows, len(all_rows))
    sample_rows = [[str(c) if c is not None else "" for c in row] for row in all_rows[sample_start:sample_end]]

    return {
        "headers": headers,
        "sample_rows": sample_rows,
        "header_row_index": header_row_index,
        "total_rows": max(total_rows - header_row_index - 1, 0),
    }


def _read_csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def _read_excel_rows(content: bytes) -> list[list[str]]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return rows


def _parse_csv_data(
    content: bytes,
    header_row_index: int,
    row_limit: int = 0,
) -> tuple[list[str], list[list[str]]]:
    all_rows = _read_csv_rows(content)
    if header_row_index >= len(all_rows):
        return [], []
    headers = [str(c).strip() for c in all_rows[header_row_index]]
    data_start = header_row_index + 1
    data_rows = all_rows[data_start:]
    if row_limit and row_limit < len(data_rows):
        data_rows = data_rows[:row_limit]
    return headers, data_rows


def _parse_excel_data(
    content: bytes,
    header_row_index: int,
    row_limit: int = 0,
) -> tuple[list[str], list[list[str]]]:
    all_rows = _read_excel_rows(content)
    if header_row_index >= len(all_rows):
        return [], []
    headers = [str(c).strip() for c in all_rows[header_row_index]]
    data_start = header_row_index + 1
    data_rows = all_rows[data_start:]
    if row_limit and row_limit < len(data_rows):
        data_rows = data_rows[:row_limit]
    return headers, data_rows
